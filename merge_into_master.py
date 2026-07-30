#!/usr/bin/env python3
"""
merge_into_master.py - fold crawler CSVs into Asia_Eateries_Master_List.xlsx.

    python merge_into_master.py --workbook Asia_Eateries_Master_List.xlsx \
                                --csv crawl_output/eatdrinkkl.csv \
                                --csv crawl_output/sethlui.csv

What it does
    1. reads every existing row from the Master List tab
    2. de-duplicates incoming rows on (country, normalised name)
    3. maps each row to your food-type taxonomy from dish keywords
    4. appends them as a new colour-coded source tier
    5. rebuilds the Master List and all category tabs, and refreshes the Index

Nothing is overwritten in place: pass --out to write to a new file if you would
rather diff the two first.
"""

import argparse, csv, glob, re, sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------- taxonomy
# First match wins, so put the specific patterns above the generic ones.
RULES_MY = [
 (r"bak kut teh|bakuteh|\bbkt\b",            "Bak Kut Teh (BKT)"),
 (r"chilli pan mee|pan mee|ban mian",        "Pan Mee, Ban Mian"),
 (r"wantan mee|wanton mee|wan tan mee",      "Wantan Mee"),
 (r"hakka",                                  "Hakka Noodles"),
 (r"pork noodle|pork ball|sang nyuk",        "Pork Noodles, Pork Ball Noodles"),
 (r"fish ?ball",                             "Fishball Noodles"),
 (r"char kuey teow|char koay teow|\bckt\b",  "Char Kuey Teow"),
 (r"hokkien mee",                            "Hokkien Mee"),
 (r"fried rice|nasi goreng",                 "Fried Rice"),
 (r"assam laksa|asam laksa|curry mee|curry laksa|white curry", "Curry Mee, Assam Laksa"),
 (r"sarawak laksa|kolo mee|kampua",          "Sarawak Laksa"),
 (r"sang har|big head prawn",                "Big Head Prawn Noodles (Sang Har Meen)"),
 (r"prawn mee|prawn noodle|har meen|mee yoke","Prawn Noodles (Har Meen, Mee Yoke)"),
 (r"beef noodle|niu zhap",                   "Beef Noodles (Niu Zhap Mien)"),
 (r"fish head noodle|fish head bee hoon",    "Fish Head Noodles"),
 (r"koay teow th'?ng|kway teow th'?ng",      "Koay Teow Th'ng"),
 (r"claypot",                                "Claypot"),
 (r"chicken rice|nasi ayam",                 "Chicken Rice"),
 (r"nasi lemak",                             "Nasi Lemak"),
 (r"porridge|congee|\bchuk\b",               "Congee, Porridge"),
 (r"steam(ed)? fish",                        "Steamed Fish Head"),
 (r"hotpot|steamboat|shabu",                 "Hotpot"),
 (r"seafood|crab|prawn|lobster",             "Chinese - Seafood"),
 (r"roast duck|braised duck|duck rice|roast pork|char siew|siu yuk", "Roast Duck, Teochew Stewed Duck"),
 (r"yong tau foo|ytf",                       "Yong Tau Foo"),
 (r"nyonya|peranakan|baba|malay |kelantan|padang|satay|rendang", "Peranakan & Malay"),
 (r"banana leaf|nasi kandar|indian|thosai|roti canai|briyani|biryani|chettinad|tandoor", "Indian & South Asian"),
 (r"omakase|sushi|kaiseki|yakitori",         "Japanese - Sushi Omakase"),
 (r"ramen|japanese|izakaya|teppan",          "Japanese - Modern"),
 (r"cendol|kuih|kueh|apam|apom|dessert|ice ?cream|gelato|pisang goreng|cake|patisserie|bakery|tart", "Snacks, Kueh & Desserts"),
 (r"\bbar\b|cocktail|speakeasy|whisky",      "Bars & Cocktails"),
 (r"steak|pizza|pasta|italian|french|spanish|western|burger|grill|bistro|fine dining|omakase counter|brunch|cafe|coffee|kopitiam|kopi|toast", "Modern & Western"),
]
RULES_SG = [
 (r"bak kut teh",                            "SG Bak Kut Teh"),
 (r"bak chor mee|minced (pork|meat) noodle|pork noodle", "SG Noodles - Bak Chor Mee"),
 (r"wanton mee|wonton mee|wanton noodle",    "SG Noodles - Wanton Mee"),
 (r"prawn mee|prawn noodle",                 "SG Noodles - Prawn Mee"),
 (r"laksa",                                  "SG Noodles - Laksa"),
 (r"fish ?ball",                             "SG Noodles - Fishball"),
 (r"beef noodle",                            "SG Noodles - Beef"),
 (r"ban mian|mee hoon kueh|curry chicken noodle", "SG Noodles - Curry & Ban Mian"),
 (r"kway teow|hokkien mee|hor fun",          "SG Fried Kway Teow & Hokkien Mee"),
 (r"chicken rice",                           "SG Chicken Rice"),
 (r"claypot",                                "SG Claypot Rice"),
 (r"fish soup|fish head|sliced fish",        "SG Fish Soup & Fish Head"),
 (r"herbal|mutton soup|pig'?s? organ|bak kut", "SG Soup - Herbal & Mutton"),
 (r"porridge|congee",                        "SG Congee & Porridge"),
 (r"kway chap|braised duck|duck rice|roast",  "SG Kway Chap & Braised Duck"),
 (r"yong tau foo",                           "SG Yong Tau Foo"),
 (r"carrot cake|oyster omelette|chai tow",   "SG Oyster Omelette & Carrot Cake"),
 (r"kueh|curry puff|bak chang|snack|dessert|cake|bakery|tart|ice ?cream", "SG Snacks & Kueh"),
 (r"zi char|cze char|seafood|crab",          "SG Zi Char & Cooked Food"),
 (r"nasi lemak|malay|nasi padang",           "SG Nasi Lemak & Malay"),
 (r"peranakan|nyonya",                       "SG Peranakan"),
 (r"indian|biryani|briyani|thosai|prata|banana leaf", "SG Indian"),
 (r"omakase|sushi|kaiseki|yakitori|kappo",   "Japanese - Sushi Omakase"),
 (r"ramen|japanese|izakaya",                 "Japanese - Modern"),
]
DEFAULT = {"Malaysia": "Taichow Spots", "Singapore": "SG Hawker - Other"}


def categorise(country, blob):
    low = blob.lower()
    for pat, cat in (RULES_MY if country == "Malaysia" else RULES_SG):
        if re.search(pat, low):
            return cat
    return DEFAULT.get(country, "SG Hawker - Other")


def norm(n):
    n = str(n).lower()
    n = re.sub(r"\(.*?\)", " ", n)
    n = re.sub(r"[\u4e00-\u9fff\u3040-\u30ff]", " ", n)
    n = re.sub(r"\b(restoran|restaurant|kedai|the|by|singapore|sg)\b", " ", n)
    return re.sub(r"[^a-z0-9]+", "", n)


# ---------------------------------------------------------------- routing
JP = {"Japanese - Sushi Omakase", "Japanese - Sushi Kappo", "Japanese - Kappo",
      "Japanese - Kaiseki", "Japanese - Modern", "Japanese - Yakitori", "Japanese - Multi-concept"}
SG_MAP = {
 "SG Noodles - Bak Chor Mee": "SG · Noodles & Hawker", "SG Noodles - Wanton Mee": "SG · Noodles & Hawker",
 "SG Noodles - Prawn Mee": "SG · Noodles & Hawker", "SG Noodles - Laksa": "SG · Noodles & Hawker",
 "SG Noodles - Fishball": "SG · Noodles & Hawker", "SG Noodles - Beef": "SG · Noodles & Hawker",
 "SG Noodles - Curry & Ban Mian": "SG · Noodles & Hawker",
 "SG Fried Kway Teow & Hokkien Mee": "SG · Noodles & Hawker", "SG Hawker - Other": "SG · Noodles & Hawker",
 "SG Chicken Rice": "SG · Rice, Soup & Porridge", "SG Claypot Rice": "SG · Rice, Soup & Porridge",
 "SG Bak Kut Teh": "SG · Rice, Soup & Porridge", "SG Fish Soup & Fish Head": "SG · Rice, Soup & Porridge",
 "SG Soup - Herbal & Mutton": "SG · Rice, Soup & Porridge", "SG Congee & Porridge": "SG · Rice, Soup & Porridge",
 "SG Kway Chap & Braised Duck": "SG · Rice, Soup & Porridge", "SG Yong Tau Foo": "SG · Rice, Soup & Porridge",
 "SG Oyster Omelette & Carrot Cake": "SG · Snacks, Kueh & Sides", "SG Snacks & Kueh": "SG · Snacks, Kueh & Sides",
 "SG Zi Char & Cooked Food": "SG · Zi Char & Seafood", "SG Seafood": "SG · Zi Char & Seafood",
 "SG Indian": "SG · Indian, Malay & Peranakan", "SG Nasi Lemak & Malay": "SG · Indian, Malay & Peranakan",
 "SG Peranakan": "SG · Indian, Malay & Peranakan", "SG International & Western": "SG · International & Western"}
MY_MAP = {
 "Wantan Mee": "MY · Noodles Egg & Pan Mee", "Pan Mee, Ban Mian": "MY · Noodles Egg & Pan Mee",
 "Hakka Noodles": "MY · Noodles Egg & Pan Mee", "Fishball Noodles": "MY · Noodles Egg & Pan Mee",
 "Pork Noodles, Pork Ball Noodles": "MY · Noodles Egg & Pan Mee",
 "Drunken Noodles, Rice Wine Noodles": "MY · Noodles Egg & Pan Mee",
 "Noodles - Pan Mee": "MY · Noodles Egg & Pan Mee", "Noodles - Chilli Pan Mee": "MY · Noodles Egg & Pan Mee",
 "Char Kuey Teow": "MY · Noodles Fried & Wok", "Hokkien Mee": "MY · Noodles Fried & Wok",
 "Fried Rice": "MY · Noodles Fried & Wok",
 "Curry Mee, Assam Laksa": "MY · Noodles Soup & Laksa", "Sarawak Laksa": "MY · Noodles Soup & Laksa",
 "Noodles - Sarawak": "MY · Noodles Soup & Laksa", "Koay Teow Th'ng": "MY · Noodles Soup & Laksa",
 "Noodles - Seafood & Lala": "MY · Noodles Soup & Laksa",
 "Prawn Noodles (Har Meen, Mee Yoke)": "MY · Noodles Soup & Laksa",
 "Big Head Prawn Noodles (Sang Har Meen)": "MY · Noodles Soup & Laksa",
 "Beef Noodles (Niu Zhap Mien)": "MY · Noodles Soup & Laksa", "Fish Head Noodles": "MY · Noodles Soup & Laksa",
 "Chicken Rice": "MY · Rice, Claypot & Congee", "Claypot": "MY · Rice, Claypot & Congee",
 "Nasi Lemak": "MY · Rice, Claypot & Congee", "Congee, Porridge": "MY · Rice, Claypot & Congee",
 "Bak Kut Teh (BKT)": "MY · Bak Kut Teh", "Steamed Fish Head": "MY · Seafood & Hotpot",
 "Hotpot": "MY · Seafood & Hotpot", "Chinese - Seafood": "MY · Seafood & Hotpot",
 "Roast Duck, Teochew Stewed Duck": "MY · Roast Duck & Yong Tau Foo",
 "Yong Tau Foo": "MY · Roast Duck & Yong Tau Foo",
 "Taichow Spots": "MY · Tai Chow & Kopitiam", "KSHF (Kopitiam / Street Hawker Fare)": "MY · Tai Chow & Kopitiam",
 "Peranakan & Malay": "MY · Peranakan & Malay", "Indian & South Asian": "MY · Indian & South Asian",
 "Snacks, Kueh & Desserts": "MY · Snacks, Kueh & Desserts", "Bars & Cocktails": "MY · Bars & Cocktails",
 "Modern & Western": "MY · Modern & Western", "Steakhouse / Wine Bar": "MY · Modern & Western"}

TAB_ORDER = [
 ("SG · Omakase & Japanese", "Sushi, kappo, kaiseki, yakitori and modern Japanese counters."),
 ("SG · Noodles & Hawker", "Bak chor mee, wanton, prawn mee, laksa, fishball, beef, fried kway teow, Hokkien mee."),
 ("SG · Rice, Soup & Porridge", "Chicken rice, claypot rice, bak kut teh, fish soup, soups, congee, kway chap, yong tau foo."),
 ("SG · Zi Char & Seafood", "Cooked food / zi char, Chinese restaurants and seafood."),
 ("SG · Indian, Malay & Peranakan", "Indian, biryani, nasi lemak, Malay and Peranakan."),
 ("SG · Snacks, Kueh & Sides", "Oyster omelette, carrot cake, curry puffs, kueh, patisserie, tea."),
 ("SG · International & Western", "Italian, Spanish, French, Thai, Korean, steakhouses, modern Singaporean."),
 ("MY · Fine Dining & Japanese", "Omakase counters across KL, Penang and JB."),
 ("MY · Modern & Western", "MICHELIN-starred and Selected modern Malaysian, European and steakhouses."),
 ("MY · Noodles Egg & Pan Mee", "Wantan mee, pan mee / ban mian, pork noodles, hakka, fishball."),
 ("MY · Noodles Fried & Wok", "Char kuey teow, Hokkien mee, fried rice."),
 ("MY · Noodles Soup & Laksa", "Curry mee, assam and Sarawak laksa, prawn, beef, fish head, koay teow th'ng."),
 ("MY · Rice, Claypot & Congee", "Chicken rice, claypot, nasi lemak, congee."),
 ("MY · Bak Kut Teh", "Bak kut teh, dry and soup."),
 ("MY · Seafood & Hotpot", "Steamed fish head, hotpot, seafood restaurants."),
 ("MY · Roast Duck & Yong Tau Foo", "Roast and Teochew stewed duck, yong tau foo."),
 ("MY · Peranakan & Malay", "Nyonya, Baba and Malay kitchens."),
 ("MY · Indian & South Asian", "Indian, Sri Lankan and South Asian."),
 ("MY · Tai Chow & Kopitiam", "Tai chow spots and kopitiam / street hawker fare."),
 ("MY · Snacks, Kueh & Desserts", "Nyonya kueh, apom, pancakes and sweets."),
 ("MY · Bars & Cocktails", "Asia's 50 Best Bars entries."),
 ("TH · Bangkok", "Bangkok noodles, Isaan, Thai-Chinese, seafood, bars and dessert."),
 ("Other · Travel & Unplaced", "Entries whose city could not be confirmed."),
]


def route(country, cat):
    if country == "Other":
        return "Other · Travel & Unplaced"
    if country == "Thailand":
        return "TH · Bangkok"
    if country == "Singapore":
        return "SG · Omakase & Japanese" if cat in JP else SG_MAP.get(cat, "SG · International & Western")
    if cat in JP:
        return "MY · Fine Dining & Japanese"
    return MY_MAP.get(cat, "MY · Tai Chow & Kopitiam")


# ---------------------------------------------------------------- styling
FONT = "Arial"
HF = Font(name=FONT, bold=True, size=10, color="FFFFFF")
HFILL = PatternFill("solid", fgColor="1F3B4D")
BF = Font(name=FONT, size=10)
THIN = Side(style="thin", color="C9CDD1")
BD = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
FILL = {
 "Curated (verified)":                 PatternFill("solid", fgColor="EAF1F6"),
 "MICHELIN / web guides (Jul 2026)":   PatternFill("solid", fgColor="FBF0DC"),
 "Omakase Diaries (Google Drive)":     PatternFill("solid", fgColor="EDE7F6"),
 "Chiefeater / Bangsar Babe (Jul 2026)": PatternFill("solid", fgColor="E3F2E1"),
 "SETHLUI.com (Jul 2026)":             PatternFill("solid", fgColor="FDE7EF"),
 "Blog crawl":                         PatternFill("solid", fgColor="FFF3E0"),
 "Malaysia Chiak (Mar 2024)":          PatternFill("solid", fgColor="F7F7F5"),
}
MASTER_H = ["#", "Source", "Country", "City / State", "Food Type Category", "Cuisine / Style", "Name",
            "Area / Location", "Address", "Phone", "Typical Hours", "Accolades", "Price Guide (per pax)",
            "Instagram / Web", "What To Order / Signature", "Google Rating", "Notes"]
MASTER_W = [5, 28, 11, 18, 34, 30, 42, 28, 50, 18, 30, 42, 26, 30, 44, 9, 50]
TAB_H = ["#", "Food Type Category", "Name", "City / State", "Area / Location", "Address", "Phone",
         "Typical Hours", "Accolades", "Price Guide (per pax)", "Instagram / Web",
         "What To Order / Signature", "Google Rating", "Source", "Notes"]
TAB_W = [5, 34, 42, 18, 26, 50, 18, 30, 42, 26, 30, 44, 9, 28, 50]


def write_sheet(ws, headers, widths, rows, rating_col):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font, cell.fill, cell.border = HF, HFILL, BD
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for i, (src, vals) in enumerate(rows, 1):
        for c, v in enumerate([i] + vals, 1):
            cell = ws.cell(row=i + 1, column=c, value=v)
            cell.font, cell.border = BF, BD
            cell.fill = FILL.get(src, FILL["Blog crawl"])
            cell.alignment = (Alignment(horizontal="center", vertical="top")
                              if c in (1, rating_col) else Alignment(vertical="top", wrap_text=True))
        ws.cell(row=i + 1, column=rating_col).number_format = "0.0"
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[1].height = 32
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows)+1}"


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--workbook", required=True)
    ap.add_argument("--csv", action="append", required=True, help="crawler CSV (repeatable, globs ok)")
    ap.add_argument("--out", help="output path (default: overwrite --workbook)")
    ap.add_argument("--source-label", default="Blog crawl")
    ap.add_argument("--dry-run", action="store_true")
    a = ap.parse_args()

    wb0 = openpyxl.load_workbook(a.workbook, data_only=True)
    ALL, seen = [], set()
    for r in wb0["Master List"].iter_rows(min_row=2, values_only=True):
        if r[0] is None:
            continue
        row = list(r[2:17])
        ALL.append((r[1], row))
        seen.add((row[0], norm(row[4])))
    print(f"existing rows: {len(ALL)}")

    paths = [p for pat in a.csv for p in sorted(glob.glob(pat))]
    if not paths:
        sys.exit("no CSVs matched")

    added = skipped = noname = 0
    for path in paths:
        with open(path, newline="", encoding="utf-8") as fh:
            for rec in csv.DictReader(fh):
                name = (rec.get("name") or "").strip()
                country = (rec.get("country") or "").strip()
                if not name or country not in ("Malaysia", "Singapore"):
                    noname += 1
                    continue
                if (rec.get("closed") or "").strip().lower() == "yes":
                    continue
                key = (country, norm(name))
                if key in seen:
                    skipped += 1
                    continue
                seen.add(key)
                blob = " ".join(filter(None, [rec.get("post_title"), name, rec.get("address")]))
                cat = categorise(country, blob)
                city = "Singapore" if country == "Singapore" else "(state not stated)"
                if country == "Malaysia":
                    for st in ("Kuala Lumpur", "Selangor", "Penang", "Johor", "Perak", "Melaka",
                               "Negeri Sembilan", "Kedah", "Pahang", "Sabah", "Sarawak"):
                        if st.lower() in (rec.get("address") or "").lower():
                            city = st
                            break
                ALL.append((a.source_label, [
                    country, city, cat,
                    "Hawker / restaurant", name, rec.get("area") or "-",
                    rec.get("address") or "-", rec.get("phone") or "-", rec.get("hours") or "-",
                    "-", rec.get("price") or "-", rec.get("url") or "-", "-", None,
                    f"Crawled from {rec.get('source','blog')} ({rec.get('date','')}). "
                    f"Category inferred from the post title; details as published, not independently verified."]))
                added += 1

    print(f"added {added} | duplicates skipped {skipped} | unusable rows {noname}")
    if a.dry_run:
        return

    buckets = {n: [] for n, _ in TAB_ORDER}
    for s, r in ALL:
        buckets[route(r[0], r[2])].append((s, r))
    for k in buckets:
        buckets[k].sort(key=lambda x: (x[1][2], x[1][4]))

    wb = openpyxl.Workbook()
    idx = wb.active
    idx.title = "Index"
    idx["A1"] = "Asia Eateries - Singapore, Malaysia, Thailand"
    idx["A1"].font = Font(name=FONT, bold=True, size=15)
    for c, h in enumerate(["Tab", "What's in it", "Records"], 1):
        cell = idx.cell(row=3, column=c, value=h)
        cell.font, cell.fill, cell.border = HF, HFILL, BD
        cell.alignment = Alignment(horizontal="center", vertical="center")
    cell = idx.cell(row=4, column=1, value="Master List")
    cell.font = Font(name=FONT, size=10, bold=True, color="0563C1", underline="single")
    cell.hyperlink, cell.border = "#'Master List'!A1", BD
    d = idx.cell(row=4, column=2, value="Everything in one filterable table.")
    d.font, d.border = BF, BD
    f = idx.cell(row=4, column=3, value=f"=COUNTA('Master List'!$G$2:$G${len(ALL)+1})")
    f.font, f.border, f.alignment = BF, BD, Alignment(horizontal="center")
    for i, (name, desc) in enumerate(TAB_ORDER):
        row = 5 + i
        c1 = idx.cell(row=row, column=1, value=name)
        c1.font = Font(name=FONT, size=10, color="0563C1", underline="single")
        c1.hyperlink, c1.border = f"#'{name}'!A1", BD
        c2 = idx.cell(row=row, column=2, value=desc)
        c2.font, c2.border = BF, BD
        c2.alignment = Alignment(wrap_text=True, vertical="top")
        n = len(buckets[name])
        c3 = idx.cell(row=row, column=3, value=f"=COUNTA('{name}'!$C$2:$C${n+1})")
        c3.font, c3.border, c3.alignment = BF, BD, Alignment(horizontal="center")
    tr = 5 + len(TAB_ORDER)
    idx.cell(row=tr, column=1, value="TOTAL (tabs)").font = Font(name=FONT, bold=True, size=10)
    t = idx.cell(row=tr, column=3, value=f"=SUM(C5:C{tr-1})")
    t.font, t.alignment = Font(name=FONT, bold=True, size=10), Alignment(horizontal="center")
    idx.column_dimensions["A"].width = 34
    idx.column_dimensions["B"].width = 82
    idx.column_dimensions["C"].width = 11

    ms = wb.create_sheet("Master List")
    write_sheet(ms, MASTER_H, MASTER_W, [(s, [s] + list(v)) for s, v in ALL], 16)
    ms.freeze_panes = "C2"
    for name, _ in TAB_ORDER:
        ws = wb.create_sheet(name)
        rows = [(s, [v[2], v[4], v[1], v[5], v[6], v[7], v[8], v[9], v[10], v[11], v[12], v[13], s, v[14]])
                for s, v in buckets[name]]
        write_sheet(ws, TAB_H, TAB_W, rows, 13)
        ws.freeze_panes = "D2"

    out = a.out or a.workbook
    wb.save(out)
    print(f"wrote {out} - {len(ALL)} rows across {len(TAB_ORDER)} tabs")
    print("NOTE: formulas need one recalculation - open in Excel, or run LibreOffice headless.")


if __name__ == "__main__":
    main()
